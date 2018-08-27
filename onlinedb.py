import click
import openpyxl
import MySQLdb
import os,django
os.environ.setdefault('DJANGO_SETTINGS_MODULE',"online.settings")
django.setup()

from onlineapp.models import *


@click.group()
def cli():
    pass


@cli.command()
@click.argument('files', nargs=2)
def populatedb(files):
    # for College
    wb = openpyxl.load_workbook(files[0])
    Sheet = wb['Colleges']
    mrow = Sheet.max_row
    for row_id in range(1, mrow):
        l = list(Sheet.rows)[row_id]
        college_obj = College()
        college_obj.name = str(l[0].value)
        college_obj.acronym = str(l[1].value)
        college_obj.location = str(l[2].value)
        college_obj.contact = str(l[3].value)
        college_obj.save()

    # for student
    wb = openpyxl.load_workbook(files[0])
    Sheet = wb['Current']
    mrow = Sheet.max_row
    for row_id in range(1, mrow):
        l = list(Sheet.rows)[row_id]
        stud_obj = Student()
        stud_obj.name = str(l[0].value)
        #stud_obj.college = l[1]
        stud_obj.email = str(l[2].value)
        stud_obj.db_folder = str(l[3].value).lower()
        stud_obj.college=College.objects.get(acronym=str(l[1].value))
        stud_obj.save()

    wb=openpyxl.load_workbook(files[0])
    Sheet=wb['Deletions']
    mrow=Sheet.max_row
    for row_id in range(1,mrow):
        l = list(Sheet.rows)[row_id]
        stud_obj = Student()
        stud_obj.name = str(l[0].value)
        #stud_obj.college = l[1]
        stud_obj.email = str(l[2].value)
        stud_obj.db_folder = str(l[3].value).lower()
        stud_obj.dropped_out=True
        stud_obj.college=College.objects.get(acronym=str(l[1].value))
        stud_obj.save()

    wb=openpyxl.load_workbook(files[1])
    Sheet=wb['Sheet']
    mrow=Sheet.max_row
    for row_id in range(1,mrow):
        l=list(Sheet.rows)[row_id]
        marks_obj = MockTest1()
        try:
            marks_obj.student = Student.objects.get(db_folder=str(l[0].value.split('_')[2]).lower())
            marks_obj.problem1 = str(l[1].value)
            marks_obj.problem2 = str(l[2].value)
            marks_obj.problem3 = str(l[3].value)
            marks_obj.problem4 = str(l[4].value)
            marks_obj.total = str(l[5].value)
            marks_obj.save()
        except Exception:
            pass

@cli.command()
@click.argument('dbname',nargs=1)
def createdb(dbname):
    conn=MySQLdb.connect('localhost','root','',dbname[0])
    cur=conn.cursor()
    cur.execute("create database {}".format(dbname[0]))
    cur.close()


@cli.command()
@click.argument('dbname', nargs=1)
def dropdb(dbname):
    conn = MySQLdb.connect('localhost', 'root', '', dbname[0])
    cur = conn.cursor()
    cur.execute("drop database {}".format(dbname[0]))
    cur.close()


if __name__=='__main__':
    cli()